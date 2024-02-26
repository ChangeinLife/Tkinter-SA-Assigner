from tkinter import *
import tkinter.messagebox
from openpyxl.styles import Font, PatternFill 
import openpyxl, xlrd
from openpyxl import Workbook
import pathlib
import random
import pandas as pd

root = Tk()
root.title("Clinical Question Short Answer Assigner App") 
root.geometry('1300x800')

root.configure (bg ='#e69c9c')
root.config(highlightbackground="black", highlightthickness=5)



file = pathlib.Path("BioClassSANAssigner.xlsx")

if file.exists():
    pass
else:
    file=Workbook()
    sheet=file.active
    sheet['A1']='Student Name'
    sheet['A1'].font = Font(size = 14, bold = True)
    sheet['A1'].fill = PatternFill(fill_type='solid', start_color = "FFFF00", end_color= 'FFFF00')
    sheet['B1']='Student PIN'
    sheet['B1'].font = Font(size = 14, bold = True)
    sheet['B1'].fill = PatternFill(fill_type='solid', start_color = "FFFF00", end_color= 'FFFF00')
    sheet['C1']='CW #1 SA'
    sheet['C1'].font = Font(size = 14, bold = True)
    sheet['C1'].fill = PatternFill(fill_type='solid', start_color = "FFFF00", end_color= 'FFFF00')

    file.save("BioClassSANAssigner.xlsx")

    
number =[]
Codes =[]

def submit(*args):
    global submit
    df = pd.read_excel("BioClass.xlsx")
    print(df["Student PIN"].values)
    Codes = (df["Student PIN"].values)

    temp =[]
    roster =[]
    added = name.get()
    name.delete(0, 'end')
    temp.append(added)
    print(temp)

    froster = [item.strip().lower() for item in temp]
    y = list(froster)
    print(y)

    file = openpyxl.load_workbook("BioClassSANAssigner.xlsx")
    sheet=file.active
    for i, value in enumerate(y, start=2):
            sheet.cell(column =1, row=sheet.max_row+1).value = value

    file.save("BioClassSANAssigner.xlsx")

    number =[]
    entered = int(pincode.get())
    pincode.delete(0, 'end')
    number.append(entered)
    print(number)

    file = openpyxl.load_workbook("BioClassSANAssigner.xlsx")
    sheet=file.active
    for i, value in enumerate(number, start=2):
            sheet.cell(column =2, row=sheet.max_row).value = value

    file.save("BioClassSANAssigner.xlsx")

    sanumber = random.randint(1,4)
    print(sanumber)
   
    
    if entered in Codes:
            tkinter.messagebox.showinfo ("SA Question #", f"Your CW SA Question number is {sanumber}")
    else: 
        tkinter.messagebox.showinfo ("SA Question #", f"Your PIN is incorrect! See Instructor")
        sanumber = "default"  
   
    w = sanumber
    file = openpyxl.load_workbook("BioClassSANAssigner.xlsx")
    sheet=file.active
    sheet.cell (column=3,row=sheet.max_row,value=w)
    file.save("BioClassSANAssigner.xlsx")
         

global mylabel
mylabel = Label(root, font=("Arial", 25), text= "CW Short Answer Assigner")
mylabel.pack(pady = 20, ipady = 2)

global mylabel2
mylabel2 = Label(root, font=("Arial", 25), text= "INSTRUCTIONS:")
mylabel2.pack(pady=20, ipady= 2)

global mylabel3
mylabel3 = Label(root, font=("Arial", 25), text= "1) Enter Your OFFICIAL NAME (format: last name (space) first name, LOWERCASE ONLY)")
mylabel3.pack(pady=20, ipady= 2)

global name
name = Entry(root, font=("Arial", 25), width = 75, borderwidth= 10)
name.bind("<Return>", submit)
name.pack(pady=20, ipady= 2)

global mylabel4
mylabel4 = Label(root, font=("Arial", 25), text= "2) Enter Your 4 DIGIT PIN.")
mylabel4.pack(pady=20, ipady= 1)

global pin_number
pincode = Entry(root, font=("Arial", 25), width = 75, borderwidth= 10)
pincode.bind("<Return>", submit)
pincode.pack(pady=20, ipady= 2)

global mylabel5
mylabel5 = Label(root, font=("Arial", 25), text= "3) PRESS 'SUBMIT' OR ENTER")
mylabel5.pack(pady=20, ipady= 1)

global submitButton
submitButton = Button(root, font=('Arial', 25), text= 'SUBMIT', fg = 'blue', command = submit)
submitButton.pack(pady=10, ipady= 2)

root.mainloop()